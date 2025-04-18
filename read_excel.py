import os

import os, re
import os.path as path
import pandas as pd
import shutil
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from utility import *

def resize(filename):
    wb = load_workbook(filename)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        df = pd.read_excel(filename, sheet).fillna('-')
        df.loc[len(df)] = list(df.columns)
        for col in df.columns:
            index = list(df.columns).index(col)
            letter = get_column_letter(index + 1)
            collen = df[col].apply(lambda x: len(str(x).encode())).max()
            ws.column_dimensions[letter].width = collen * 1 + 4
    wb.save(filename)

def findFile(pattern, base='.'):
    regex = re.compile(pattern)
    matches = []
    for root, dirs, files in os.walk(base):
        for f in files:
            if regex.match(f):
                matches.append(path.join(root, f))
    return matches

def GetFiles(path_k):
    paths = os.walk(path_k)
    files = []
    for path, dir_lst, file_lst in paths:
        # print("dir_cnt:", len(dir_lst))
        # print("file_cnt:", len(file_lst))
        for dir_name in file_lst:
            f = os.path.join(path, dir_name)
            files.append(f)
        return files
    return files

def changeExtName():
    path = "D:\\work\\dexuan\\2501\\play_test\\audio_rename"
    srcFiles = GetDirs(path)
    for i, val in enumerate(srcFiles):
        fullName = val["fullname"]
        cnt = change_specific_extensions(fullName, ".mp3", ".aac")
        print(f"fullname:{fullName}, cnt:{cnt}")

def PrintFile(fileName):
    with open(fileName, "r", encoding='utf-8') as file:
        content = file.read()
        print("fileName:", fileName)
        print(content)

def procData(excelFile, langDstPath, audioSrcPath, column):
    print("excelFile:", excelFile, " langDstPath:", langDstPath, " audioSrcPath:", audioSrcPath, " column:", column)

    if not os.path.exists(langDstPath):
        os.makedirs(langDstPath)  # 创建路径

    # slogan
    sloganPath = langDstPath + "\\slogan"
    if not os.path.exists(sloganPath):
        os.makedirs(sloganPath)  # 创建路径
    killSloganPath = sloganPath + "\\kill"
    deadSloganPath = sloganPath + "\\dead"
    if not os.path.exists(killSloganPath):
        os.makedirs(killSloganPath)  # 创建路径
    if not os.path.exists(deadSloganPath):
        os.makedirs(deadSloganPath)  # 创建路径

    killSloganFile = killSloganPath + "\\" + "slogan.dat"
    deadSloganFile = deadSloganPath + "\\" + "slogan.dat"

    # audio
    audioPath = langDstPath + "\\audio"
    if not os.path.exists(audioPath):
        os.makedirs(audioPath)  # 创建路径
    killAudioPath = audioPath + "\\kill"
    deadAudioPath = audioPath + "\\dead"
    if not os.path.exists(killAudioPath):
        os.makedirs(killAudioPath)  # 创建路径
    if not os.path.exists(deadAudioPath):
        os.makedirs(deadAudioPath)  # 创建路径

    df = pd.read_excel(excelFile)
    pageSize = 20
    with open(killSloganFile, "w", encoding='utf-8') as kf:
        killIdx = 1
        for i in range(killIdx, killIdx+pageSize):
            slogan = str(df.iloc[i, column])
            slogan = discard_end_char(slogan)
            audioFileName = find_file_containing_ignore_case(audioSrcPath, slogan)
            # audioFiles = findFile(audioFileName, audioSrcPath)
            if(len(audioFileName) == 0):
                print(f"find kill audio file failed, idx:{i}, path:{audioSrcPath}, slogan:{slogan}")
                continue
            num_acc = f"{i:0>2}.aac"
            srcFile = audioSrcPath + "\\" + audioFileName
            dstFile = killAudioPath + "\\" + num_acc
            # print("move kill src:", srcFile, " dst:", dstFile)
            try:
                shutil.move(srcFile, dstFile)
            except FileNotFoundError as e:
                print(f"kill failed 文件未找到错误: {e}, srcFile:{srcFile}, dstFile:{dstFile}")
                continue

            slogan += "\n"
            kf.write(slogan)

    with open(deadSloganFile, "w", encoding='utf-8') as deadSloganFile:
        deadIdx = 22
        for i in range(deadIdx, deadIdx + pageSize):
            #print("i:", i, "column:", column)
            slogan = str(df.iloc[i, column])
            slogan = discard_end_char(slogan)
            audioFileName = find_file_containing_ignore_case(audioSrcPath, slogan)
            if(len(audioFileName) == 0):
                print(f"find dead audio file failed,  idx:{i}, path:{audioSrcPath}, slogan:{slogan}" )
                continue
            num_acc = f"{i-deadIdx+1:0>2}.aac"
            srcFile = audioSrcPath + "\\" + audioFileName
            dstFile = deadAudioPath + "\\" + num_acc
            # print("move dead src:", srcFile, " dst:", dstFile)
            try:
                shutil.move(srcFile, dstFile)
            except FileNotFoundError as e:
                print(f"dead failed 文件未找到错误: {e}, srcFile:{srcFile}, dstFile:{dstFile}")
                continue
            slogan += "\n"
            deadSloganFile.write(slogan)
    pass
def GetCellData(fileName, row, column):
    #fileName = "D:\\work\\dexuan\\2501\\play_2501.xlsx"
    df = pd.read_excel(fileName)
    cell_data = df.iloc[row, column]  # 读取第row行，第column列的数据
    return cell_data

# test append string
def appendString(fileName, info):
    with open(fileName, "w") as file:
        file.write(info)
        file.write("\n")
        file.write("33333")
        file.write("\n")
        file.write("4444")
        file.write("\n")

def ExcelToLangFile(excelFile, dstPath, audioPath, column):
    if not os.path.exists(dstPath):
        os.makedirs(dstPath)  # 创建路径

    procData(excelFile, dstPath, audioPath, column)
    pass

def count_lines_in_file(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        return sum(1 for line in file)

    return 0

def checkLangCompletion(langDstPath):
    # slogan
    sloganPath = langDstPath + "\\slogan"
    killSloganPath = sloganPath + "\\kill"
    deadSloganPath = sloganPath + "\\dead"
    killSloganFile = killSloganPath + "\\" + "slogan.dat"
    deadSloganFile = deadSloganPath + "\\" + "slogan.dat"

    killSloganCnt = count_lines_in_file(killSloganFile)
    if(killSloganCnt != 20):
        print("kill slogan failed:", killSloganFile, " cnt:", killSloganCnt)
    deadSloganCnt = count_lines_in_file(deadSloganFile)
    if(deadSloganCnt != 20):
        print("dead slogan failed:", deadSloganFile, " cnt:", deadSloganCnt)


    # print text result
    PrintFile(killSloganFile)
    PrintFile(deadSloganFile)

    # audio
    audioPath = langDstPath + "\\audio"
    killAudioPath = audioPath + "\\kill"
    deadAudioPath = audioPath + "\\dead"
    kills = GetFiles(killAudioPath)
    deads = GetFiles(deadAudioPath)
    if(len(kills) != 20):
        print("kill video failed, ", killAudioPath, " cnt:", len(kills))
    if(len(deads) != 20):
        print("dead video failed:", deadAudioPath, " cnt:", len(deads))

    # print audio result
    for key, val in enumerate(kills):
        print("audio:", val)
    for key, val in enumerate(deads):
        print("audio:", val)
    pass
def checkDst(dstPath, LangtoColumn):
    for i, (lang, value) in enumerate(LangtoColumn.items()):
        langDstPath = dstPath + "\\" + lang
        checkLangCompletion(langDstPath)
        # print("finish check:", langDstPath)

    print("finish check dstPath:", dstPath)
    pass

if __name__ == '__main__':
    playDst = "D:\\work\\dexuan\\2501\\play_test\\"

    excelFile = playDst + "拯救姬话术2501.xlsx"
    audioPath = playDst + "audio_short"
    dstPath = playDst + "dstPath"

    #test
    #procData(fileName, "d:\\tmp\\kill.txt", "d:\\tmp\\dead.txt", 6)
    #data = GetCellData(fileName, 13, 3)
    #print("data:", data)

    LangtoColumn = {

        # old
        # "German": [4, "德语"],
        # "French": [5, "法语"],
        # "Polish": [6, "波兰语"],
        # "Jap": [7, "日本语"],
        # "Span": [8, "西班牙语"],
        # "Itali": [9, "意大利语"],
        # "Arabic": [10, "阿拉伯语"],
        # "CnTradition": [11, "繁体中文"],
        # "PortugueseEurope": [12, "葡萄牙语"],
        # "Korean": [13, "韩语"],
        # "Norwegian": [14, "挪威语"],
        # "Hung": [15, "匈牙利语"],
        # "Czech": [16, "捷克语"],
        # "Slovak": [17, "斯洛伐克语"],
        # "Romanian": [18, "罗马尼亚语"],
        # "Dutch": [19, "荷兰语"],
        # "Swedish": [20, "瑞典语"],
        # "Croatian": [21, "克罗地亚语"],
        # "Finnish": [22, "芬兰语"],
        # "Turkish": [23, "土耳其语"],
        # "Ukra": [24, "乌克兰语"],
        # "Danish": [25, "丹麦语"],
        # "PortugueseBrazil": [26, "巴西葡萄牙语"],
        # "Greek": [27, "希腊语"],
        # "Russian": [28, "俄语"],



        "cn": [2, "中文"],
        "en": [3, "英语"],
        "de": [4, "德语"],
        "fr": [5, "法语"],
        "pl": [6, "波兰语"],
        "jp": [7, "日本语"],
        "es": [8, "西班牙语"],
        "it": [9, "意大利语"],
        "ar": [10, "阿拉伯语"],
        "tc": [11, "繁体中文"],
        "pt": [12, "葡萄牙语"],
        "kr": [13, "韩语"],
        "no": [14, "挪威语"],
        "hu": [15, "匈牙利语"],
        "cs": [16, "捷克语"],
        "sk": [17, "斯洛伐克语"],
        "ro": [18, "罗马尼亚语"],
        "nl": [19, "荷兰语"],
        "sv": [20, "瑞典语"],
        "hr": [21, "克罗地亚语"],
        "fi": [22, "芬兰语"],
        "tr": [23, "土耳其语"],
        "ua": [24, "乌克兰语"],
        "dk": [25, "丹麦语"],
        "br": [26, "巴西葡萄牙语"],
        "gr": [27, "希腊语"],
        "ru": [28, "俄语"],
    }

    # for i, (lang, value) in enumerate(LangtoColumn.items()):
    #     audioLangPath = audioPath + "\\" + value[1]
    #     langDstPath = dstPath + "\\" + lang
    #     print("lang:", lang, " col:", value[0], " audioFile:", audioLangPath, " langDstPath:", langDstPath)
    #     procData(excelFile, langDstPath, audioLangPath, value[0])

    # dstPath = "d:\\tmp\\12"
    checkDst(dstPath, LangtoColumn)
    # PrintFile("D:\\tmp\\slogan123.dat")

